import os
import argparse

from collections import defaultdict
import openpyxl

from openpyxl.styles import PatternFill, Color, Border, Side

import sys

# parser = argparse.ArgumentParser()
# parser.add_argument('-p', '--path', type=str, help='path of target directory')
# args = parser.parse_args()

############################
##### 1. base_path 설정 #####
############################
base_path = '/Users/jihyunlee/Desktop/bleeding_data_preprosessing/'

def float_to_int(x):
    '''
        일부 location cell 이 소수점으로 이뤄져있어 정수로 변환
    '''
    if (x - int(x) < 0.5):
        return int(x) 
    return int(x) + 1

def modified_list(original_list):
    '''
        float : 소수점 -> 정수
        str : 대소문자 -> 소문자
    '''
    return_list = []
    for i in original_list:
        if type(i) == float:
            return_list.append(float_to_int(i))
        elif type(i) == str:
            return_list.append(i.lower())
        else:
            return_list.append(i)
    return return_list

def modify(origin_list):
    '''
        공백 제거 : '' -> None
    '''
    modify_list = []
    final_modify_list = []

    # 앞 뒤 공백 제거
    for i in origin_list:
        if type(i) == str:
            i = i.strip()
            modify_list.append(i)
        else:
            modify_list.append(i)

    # 단일 공백 문자 제거
    for i in modify_list:
        if i == '':
            i = None
            final_modify_list.append(i)
        else:
            final_modify_list.append(i)

    return final_modify_list



def ab_grading_consensus():
    ###############################
    ##### 2. output_path 설정 ######
    ###############################
    output_path = base_path + 'Bleeding-Grading_Consensus_09-test/'
    if not os.path.exists(output_path):
        os.makedirs(output_path)

    ###############################
    ##### 3. target 폴더 설정 ######
    ###############################    
    for root, dirs, files in os.walk(base_path + 'Individual_to_Consensus_09'):
        files.sort(key=str.lower)
        patients_dict = defaultdict(list)

        for fname in files:
            if ('Store' in fname) or ('~' in fname):
                continue

            patients_dict[fname[10:20]].append(fname)
            patients_list = list(patients_dict.values())

            '''
            patients_dict = 
            {
                '402_ch1_01': ['01_G_01_R_402_ch1_01_Bleeding-grading-Individual_07.xlsx', '01_G_01_R_402_ch1_01_Bleeding-grading-Individual_08.xlsx'], 
                '402_ch1_03': ['01_G_01_R_402_ch1_03_Bleeding-grading-Individual_07.xlsx', '01_G_01_R_402_ch1_03_Bleeding-grading-Individual_08.xlsx'], 
                '420_ch1_01': ['01_G_01_R_420_ch1_01_Bleeding-grading-Individual_08.xlsx', '01_G_01_R_420_ch1_01_Bleeding-grading-Individual_09.xlsx'], 
                '420_ch1_03': ['01_G_01_R_420_ch1_03_Bleeding-grading-Individual_08.xlsx', '01_G_01_R_420_ch1_03_Bleeding-grading-Individual_09.xlsx'], 
                '423_ch1_01': ['01_G_01_R_423_ch1_01_Bleeding-grading-Individual_07.xlsx', '01_G_01_R_423_ch1_01_Bleeding-grading-Individual_09.xlsx'], 
                '423_ch1_03': ['01_G_01_R_423_ch1_03_Bleeding-grading-Individual_07.xlsx', '01_G_01_R_423_ch1_03_Bleeding-grading-Individual_09.xlsx'], 
                '427_ch1_01': ['01_G_01_R_427_ch1_01_Bleeding-grading-Individual_08.xlsx', '01_G_01_R_427_ch1_01_Bleeding-grading-Individual_09.xlsx'], 
                '427_ch1_03': ['01_G_01_R_427_ch1_03_Bleeding-grading-Individual_08.xlsx', '01_G_01_R_427_ch1_03_Bleeding-grading-Individual_09.xlsx']
            }

            patients_list = 
            [
                ['01_G_01_R_402_ch1_01_Bleeding-grading-Individual_07.xlsx', '01_G_01_R_402_ch1_01_Bleeding-grading-Individual_08.xlsx'], 
                ['01_G_01_R_402_ch1_03_Bleeding-grading-Individual_07.xlsx', '01_G_01_R_402_ch1_03_Bleeding-grading-Individual_08.xlsx'], 
                ['01_G_01_R_420_ch1_01_Bleeding-grading-Individual_08.xlsx', '01_G_01_R_420_ch1_01_Bleeding-grading-Individual_09.xlsx'], 
                ['01_G_01_R_420_ch1_03_Bleeding-grading-Individual_08.xlsx', '01_G_01_R_420_ch1_03_Bleeding-grading-Individual_09.xlsx'], 
                ['01_G_01_R_423_ch1_01_Bleeding-grading-Individual_07.xlsx', '01_G_01_R_423_ch1_01_Bleeding-grading-Individual_09.xlsx'], 
                ['01_G_01_R_423_ch1_03_Bleeding-grading-Individual_07.xlsx', '01_G_01_R_423_ch1_03_Bleeding-grading-Individual_09.xlsx'], 
                ['01_G_01_R_427_ch1_01_Bleeding-grading-Individual_08.xlsx', '01_G_01_R_427_ch1_01_Bleeding-grading-Individual_09.xlsx'], 
                ['01_G_01_R_427_ch1_03_Bleeding-grading-Individual_08.xlsx', '01_G_01_R_427_ch1_03_Bleeding-grading-Individual_09.xlsx']
            ]
            '''

        for patients in patients_list: # patients = ['01_G_01_R_402_ch1_01_Bleeding-grading-Individual_07.xlsx', '01_G_01_R_402_ch1_01_Bleeding-grading-Individual_08.xlsx']
            patients.sort(key=str.lower)

            # 새로운 엑셀 파일 생성
            write_wb = openpyxl.Workbook()
            filename = patients[0][:-8] + '.xlsx' # filename = 01_G_01_R_402_ch1_01_Bleeding-grading-Individual.xlsx

            # 각 파일 읽기
            full_name1 = os.path.join(root, patients[0])
            read_wb1 = openpyxl.load_workbook(full_name1, data_only=True, read_only=True)

            full_name2 = os.path.join(root, patients[1])
            read_wb2 = openpyxl.load_workbook(full_name2, data_only=True, read_only=True)

            # 첫 번째 파일 기준 - 시트 불러오기 ['ch1', 'ch3', 'ch5']
            channel = []
            for worksheet in read_wb1.worksheets: 
                channel.append(worksheet.title)
            
            # 엑셀 시트 생성
            for i in channel: # ['ch1', 'ch3', 'ch5']
                ws = write_wb.create_sheet()
                ws.title = i

                # 첫 번째 엑셀 파일
                ws1 = read_wb1[i]
                ws1_all_value = []
                for row in ws1.rows:
                    row_value = []
                    for cell in row:
                        row_value.append(cell.value)
                    ws1_all_value.append(row_value)

                # 두 번째 엑셀 파일
                ws2 = read_wb2[i]
                ws2_all_value = []
                for row in ws2.rows:
                    row_value = []
                    for cell in row:
                        row_value.append(cell.value)
                    ws2_all_value.append(row_value)

                # 엑셀 column 작성
                write_ws = write_wb[i]
                for _ in range(2):
                    write_ws.append(ws1_all_value.pop(0))
                    ws2_all_value.pop(0)

                write_ws['A1'] = 'location (x,y)'
                write_ws['H1'] = 'Timestamp'
                write_ws['J1'] = 'Site'
                write_ws['L1'] = 'Cause'
                write_ws['N1'] = 'Characteristics'
                write_ws['Q1'] = 'Identical'
                write_ws['R1'] = 'Comment'
                write_ws['S1'] = 'Remedy1'
                write_ws['V1'] = 'Remedy2'
                write_ws['Y1'] = 'Remedy 3'
                write_ws['AB1'] = 'Remedy 4'
                write_ws['AE1'] = 'Remedy 5'
                
                # 데이터 입력
                tmp1, tmp2, tmp3 = 0, 0, 0
                # 첫 번째 파일 입력
                for i, row_value in enumerate(ws1_all_value):
                    # sheet 만 존재하고 데이터 없는 경우
                    if (row_value[0] == None) or (str(row_value[0]).lower() == 'x'):
                        break
                    
                    tmp_row_value1 = row_value[9:] # ['LN3', 2, 't', None, 1, 2, None, None, None, 'b', None, 1, None, None, None, None, None, None, None, None, None, None, None, None, None]
                    tmp_row_value2 = row_value[:9] # [617, 982, 462, 736, 327, 520, 1, '00:14:02:23', '00:14:07:03']
                    
                    row_value1 = modified_list(tmp_row_value1) # ['ln3', 2, 't', None, 1, 2, None, None, None, 'b', None, 1, None, None, None, None, None, None, None, None, None, None, None, None, None]
                    row_value2 = modified_list(tmp_row_value2) # [617, 982, 462, 736, 327, 520, 1, '00:14:02:23', '00:14:07:03']
                    
                    # 데이터 입력
                    for j in range(len(row_value1)):
                        write_ws.cell(row=i+3+tmp1, column=j+10, value=row_value1[j])
                    tmp1 += 2
                    
                    # 시간 입력
                    for j in range(len(row_value2)):
                        write_ws.cell(row=i+5+tmp2, column=j+1, value=row_value2[j])
                    tmp2 += 2

                # 두 번째 파일 입력
                for i, row_value in enumerate(ws2_all_value):
                    # sheet 만 존재하고 데이터 없는 경우
                    if row_value[0] == None or (str(row_value[0]).lower() == 'x'):
                        break
                    
                    tmp_row_value1 = row_value[9:]
                    row_value1 = modified_list(tmp_row_value1)

                    for j in range(len(row_value1)):
                        write_ws.cell(row=i+4+tmp3, column=j+10, value=row_value1[j])
                    tmp3 += 2      

                write_ws.insert_cols(10)
                write_ws.cell(row=1, column=10, value='알치:0, 불일치:1')
                write_ws.cell(row=2, column=10, value='Consensus')

                # 일치 여부 확인
                row = 3
                check_index = [12, 13, 15, 16, 17, 18, 20, 22, 23, 25, 26, 28, 29, 31, 32, 34]
                '''
                consensus 를 위해 비교 항목
                check index = {
                    Site > vessel
                    Cause > cause
                    Characteristics > 1, 2, 3 column
                    Identical
                    Remedy > 1, 3 column
                }
                '''
                while write_ws.cell(row+2, 1).value != None:
                    row0 = []
                    row1 = []

                    for i in check_index:
                        row0.append(write_ws.cell(row, i).value)
                        row1.append(write_ws.cell(row+1, i).value)

                    # 공백 제거 (' ' -> None)
                    row_0 = modify(row0)
                    row_1 = modify(row1)

                    if row_0 == row_1:
                        if row0 != row1:
                            print(row, patients)

                    if (row_0 == row_1):
                        for m in range(3):
                            write_ws.cell(row=row+m, column=10, value=0)
                        for e, index in enumerate(check_index):
                            write_ws.cell(row=row+2, column=index, value=row_0[e])
                        for w in range(34):
                            write_ws.cell(row=row+2, column=w+1).fill = PatternFill('solid', fgColor='EDEBEB')
                    else:
                        for m in range(2):
                            write_ws.cell(row=row+m, column=10, value=1)

                        for w in range(34):
                            write_ws.cell(row=row+2, column=w+1).fill = PatternFill('solid', fgColor='EDEBEB')

                    row += 3
            
                # 셀 병합
                merge_list = ['A1:F1', 'H1:I1', 'K1:L1', 'M1:N1', 'O1:Q1', 'T1:V1', 'W1:Y1', 'Z1:AB1', 'AC1:AE1', 'AF1:AH1', 'A2:B2', 'C2:D2', 'E2:F2']
                for i in merge_list:
                    write_ws.merge_cells(i)
                
                # 셀 테두리
                border_list = [2, 4, 6, 7, 9, 10, 12, 14, 17, 18, 19, 22, 25, 28, 31, 34]
                horizontal_border = Border(top=Side(style='thin'), 
                                            bottom=Side(style='thin')
                                            )

                vertical_border = Border(right=Side(style='thin'))

                for i in border_list:
                    for j in range(row-1):
                        write_ws.cell(row=j+1, column=i).border = vertical_border

                for i in range(34):
                    write_ws.cell(row=1, column=i+1).font = openpyxl.styles.fonts.Font(bold=True)
                    write_ws.cell(row=1, column=i+1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                    write_ws.cell(row=2, column=i+1).border = horizontal_border

            del write_wb['Sheet']            
            write_wb.save(output_path+filename)
        


# def main():
#     target_path = args.path
#     ab_grading_consensus(target_path)

if __name__ == '__main__':
    # main()
    ab_grading_consensus()
